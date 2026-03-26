from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Case, When, Value, IntegerField
from django.http import HttpResponse
from django.conf import settings
from datetime import date, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from ..models import Project, Task
from .utils import _jdumps, _parse_hours, superuser_required


# ── Field Constructor ─────────────────────────────────────

# (label, default_on, xlsx_label, xlsx_width)
REPORT_FIELDS = {
    'project_name':      ('Название проекта',      True,  'Проект',          26),
    'project_basis':     ('Обоснование проекта',   True,  'Обоснование пр.', 36),
    'project_initiator': ('Инициатор проекта',     True,  'Инициатор пр.',   24),
    'date':              ('Дата задачи',            True,  'Дата',            13),
    'hours':             ('Время (часы)',           True,  'Часы',            9),
    'task':              ('Название задачи',        True,  'Задача',          52),
    # optional
    'status':            ('Статус',                False, 'Статус',          18),
    'task_basis':        ('Обоснование задачи',    False, 'Обоснование зад.',36),
    'initiator':         ('Инициатор задачи',      False, 'Инициатор зад.',  24),
    'start_date':        ('Дата начала',            False, 'Начало',          13),
    'due_date':          ('Срок выполнения',        False, 'Срок',            13),
    'assigned_to':       ('Исполнитель',            False, 'Исполнитель',     24),
}

REQUIRED_FIELDS = {k for k, v in REPORT_FIELDS.items() if v[1]}  # default-on = required


def _get_active_fields(request):
    """Return ordered list of active field keys from GET params."""
    active = []
    for key in REPORT_FIELDS:
        if key in REQUIRED_FIELDS:
            active.append(key)
        elif request.GET.get(f'f_{key}') == '1':
            active.append(key)
    return active


def _task_field_value(task, key, status_map):
    if key == 'project_name':
        return task.project.name
    if key == 'project_basis':
        return task.project.description or '—'
    if key == 'project_initiator':
        return task.project.initiator or '—'
    if key == 'date':
        return task.date.strftime('%d.%m.%Y')
    if key == 'hours':
        return float(task.hours) if task.hours else '—'
    if key == 'task':
        return task.task
    if key == 'status':
        return status_map.get(task.status, task.status)
    if key == 'task_basis':
        return task.basis or '—'
    if key == 'initiator':
        return task.initiator or '—'
    if key == 'start_date':
        return task.start_date.strftime('%d.%m.%Y') if task.start_date else '—'
    if key == 'due_date':
        return task.due_date.strftime('%d.%m.%Y') if task.due_date else '—'
    if key == 'assigned_to':
        try:
            return task.assigned_to.profile.display_name or task.assigned_to.username
        except Exception:
            return task.assigned_to.username if task.assigned_to else '—'
    return '—'


# ── Report view ───────────────────────────────────────────

@login_required
def report(request):
    from ..models import ProjectMember as _PM

    projects = Project.objects.filter(user=request.user)
    tasks = Task.objects.filter(
        project__user=request.user
    ).select_related('project', 'assigned_to', 'assigned_to__profile')

    today = date.today()
    date_from  = request.GET.get('date_from', today.replace(day=1).isoformat())
    date_to    = request.GET.get('date_to', today.isoformat())
    project_id = request.GET.get('project', '')
    status_filter = request.GET.get('status', '')
    group_by   = request.GET.get('group', 'project')

    if date_from:
        tasks = tasks.filter(date__gte=date_from)
    if date_to:
        tasks = tasks.filter(date__lte=date_to)
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    if project_id:
        tasks = tasks.filter(project_id=project_id)
        proj = Project.objects.filter(pk=project_id).first()
        if proj and proj.user != request.user:
            m = _PM.objects.filter(project_id=project_id, user=request.user).first()
            if m:
                tasks = tasks.filter(
                    Q(assigned_to=request.user) |
                    Q(assigned_to__isnull=True, project__user=request.user)
                )

    tasks = tasks.order_by('project__name', 'date')
    tasks_list = list(tasks)
    total_hours = sum((t.hours for t in tasks_list if t.hours), Decimal('0'))

    active_fields = _get_active_fields(request)
    status_map = dict(Task.STATUS_CHOICES)

    if request.GET.get('format') == 'xlsx':
        return _export_xlsx(tasks_list, date_from, date_to, active_fields, status_map)

    # ── Text report grouped by project ──
    text_lines = []
    current_proj = None
    for t in tasks_list:
        if t.project != current_proj:
            current_proj = t.project
            text_lines.append('')
            # Project header
            proj_header = t.project.name
            text_lines.append(proj_header)
            text_lines.append('═' * max(len(proj_header), 30))
            # Project-level fields
            if 'project_initiator' in active_fields and t.project.initiator:
                text_lines.append(f'Инициатор: {t.project.initiator}')
            if 'project_basis' in active_fields and t.project.description:
                text_lines.append(f'Обоснование: {t.project.description}')
            text_lines.append('')

        # Task line
        parts = []
        if 'date' in active_fields:
            parts.append(t.date.strftime('%d.%m.%Y'))
        if 'hours' in active_fields and t.hours:
            parts.append(f'[{t.hours}ч]')
        if 'status' in active_fields:
            parts.append(f'({status_map.get(t.status, t.status)})')

        prefix = '  ' + ' '.join(parts) + (' — ' if parts else '  ')

        if 'task' in active_fields:
            text_lines.append(prefix + t.task)

        # Sub-fields
        subs = []
        if 'initiator' in active_fields and t.initiator:
            subs.append(f'Инициатор: {t.initiator}')
        if 'task_basis' in active_fields and t.basis:
            subs.append(f'Обоснование: {t.basis}')
        if 'assigned_to' in active_fields and t.assigned_to:
            name = _task_field_value(t, 'assigned_to', status_map)
            subs.append(f'Исполнитель: {name}')
        if 'start_date' in active_fields and t.start_date:
            subs.append(f'Начало: {t.start_date.strftime("%d.%m.%Y")}')
        if 'due_date' in active_fields and t.due_date:
            subs.append(f'Срок: {t.due_date.strftime("%d.%m.%Y")}')
        for s in subs:
            text_lines.append('    ' + s)

    text_report = '\n'.join(text_lines).strip()

    # Grouped for table view
    grouped = {}
    for t in tasks_list:
        grouped.setdefault(t.project, []).append(t)

    return render(request, 'projects/report.html', {
        'projects':        projects,
        'tasks':           tasks_list,
        'grouped':         grouped,
        'group_by':        group_by,
        'date_from':       date_from,
        'date_to':         date_to,
        'status_filter':   status_filter,
        'status_choices':  Task.STATUS_CHOICES,
        'selected_project': str(project_id) if project_id else '',
        'total':           len(tasks_list),
        'total_hours':     total_hours,
        'text_report':     text_report,
        'report_fields':   REPORT_FIELDS,
        'active_fields':   set(active_fields),
        'required_fields': REQUIRED_FIELDS,
    })


def _group_tasks(tasks_list, group_by):
    grouped = {}
    if group_by == 'project':
        for t in tasks_list:
            grouped.setdefault(t.project, []).append(t)
    elif group_by == 'date':
        for t in tasks_list:
            grouped.setdefault(t.date, []).append(t)
    else:
        grouped = {'__all__': tasks_list}
    return grouped


# ── Excel export ──────────────────────────────────────────

def _export_xlsx(tasks_list, date_from, date_to, active_fields, status_map):
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Отчёт'

    COLOR_ACCENT       = '1E40AF'
    COLOR_ROW_ODD      = 'FFFFFF'
    COLOR_ROW_EVEN     = 'F8FAFF'
    COLOR_TOTAL_BG     = 'EFF6FF'
    COLOR_DONE         = '065F46'
    COLOR_OVERDUE      = '991B1B'
    COLOR_PROJ_HEADER  = 'DBEAFE'

    thin = lambda: Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    thick_bottom = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='medium', color=COLOR_ACCENT),
    )

    hfont  = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    hfill  = PatternFill(start_color=COLOR_ACCENT, end_color=COLOR_ACCENT, fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(wrap_text=True, vertical='top', horizontal='left')
    center = Alignment(vertical='top', horizontal='center')
    pfill  = PatternFill(start_color=COLOR_PROJ_HEADER, end_color=COLOR_PROJ_HEADER, fill_type='solid')

    # Build columns from active_fields
    # project_name → always first group header; skip from per-row columns if grouping
    # We'll do project group rows + task data rows
    col_fields = [f for f in active_fields if f not in ('project_name', 'project_basis', 'project_initiator')]
    # xlable, width
    def col_meta(key):
        return REPORT_FIELDS[key][2], REPORT_FIELDS[key][3]

    # Header
    ws.row_dimensions[1].height = 26
    for ci, key in enumerate(col_fields, 1):
        label, width = col_meta(key)
        c = ws.cell(row=1, column=ci, value=label)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = thick_bottom
        ws.column_dimensions[get_column_letter(ci)].width = width

    odd_fill  = PatternFill(start_color=COLOR_ROW_ODD,  end_color=COLOR_ROW_ODD,  fill_type='solid')
    even_fill = PatternFill(start_color=COLOR_ROW_EVEN, end_color=COLOR_ROW_EVEN, fill_type='solid')

    ri = 2
    total_hours = Decimal('0')
    today = date.today()

    # Group by project
    from itertools import groupby
    from operator import attrgetter
    sorted_tasks = sorted(tasks_list, key=lambda t: t.project.name)

    for proj, proj_tasks in groupby(sorted_tasks, key=lambda t: t.project):
        proj_tasks = list(proj_tasks)

        # Project header row (spans all columns)
        if col_fields:
            proj_hours = sum((t.hours for t in proj_tasks if t.hours), Decimal('0'))
            n = len(col_fields)

            proj_label_parts = [proj.name]
            if 'project_initiator' in active_fields and proj.initiator:
                proj_label_parts.append(f'Инициатор: {proj.initiator}')
            if 'project_basis' in active_fields and proj.description:
                proj_label_parts.append(f'Обоснование: {proj.description}')

            ws.row_dimensions[ri].height = 18
            c = ws.cell(row=ri, column=1, value=' · '.join(proj_label_parts))
            c.font = Font(bold=True, color=COLOR_ACCENT, size=10, name='Calibri')
            c.fill = pfill
            c.alignment = Alignment(vertical='center', horizontal='left')
            c.border = thin()
            if n > 1:
                ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=n)
            ri += 1

        for task in proj_tasks:
            h = task.hours or None
            if h:
                total_hours += h

            fill = even_fill if ri % 2 == 0 else odd_fill
            ws.row_dimensions[ri].height = 32

            for ci, key in enumerate(col_fields, 1):
                val = _task_field_value(task, key, status_map)
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill; c.border = thin()
                is_text = key in ('task', 'task_basis')
                c.alignment = left if is_text else (
                    center if key in ('date', 'start_date', 'due_date', 'hours', 'status') else left
                )
                if key == 'status':
                    if task.status == Task.STATUS_DONE:
                        c.font = Font(color=COLOR_DONE, name='Calibri')
                    elif task.due_date and task.due_date < today and task.status != Task.STATUS_DONE:
                        c.font = Font(color=COLOR_OVERDUE, bold=True, name='Calibri')
            ri += 1

    # Summary
    if col_fields:
        n = len(col_fields)
        total_fill = PatternFill(start_color=COLOR_TOTAL_BG, end_color=COLOR_TOTAL_BG, fill_type='solid')
        summary_border = Border(top=Side(style='medium', color=COLOR_ACCENT))
        ws.row_dimensions[ri].height = 22
        for ci in range(1, n + 1):
            c = ws.cell(row=ri, column=ci)
            c.fill = total_fill; c.border = summary_border

        ws.cell(row=ri, column=1, value=f'Итого задач: {len(tasks_list)}').font = Font(bold=True, color=COLOR_ACCENT, name='Calibri')
        ws.cell(row=ri, column=1).alignment = Alignment(horizontal='left', vertical='center')
        # Find hours column
        if 'hours' in col_fields:
            hci = col_fields.index('hours') + 1
            c = ws.cell(row=ri, column=hci, value=float(total_hours))
            c.font = Font(bold=True, color=COLOR_ACCENT, name='Calibri')
            c.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(col_fields))}{ri - 1}' if col_fields else None

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    period = f'{date_from}__{date_to}' if (date_from or date_to) else 'all'
    response['Content-Disposition'] = f'attachment; filename="report_{period}.xlsx"'
    wb.save(response)
    return response
