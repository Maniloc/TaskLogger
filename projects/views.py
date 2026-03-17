import json
import logging
from functools import wraps
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncMonth, TruncDate
from django.http import HttpResponse, HttpResponseForbidden
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Project, Task


class _JsonEncoder(json.JSONEncoder):
    """Handle Decimal and date types in json.dumps."""
    def default(self, obj):
        from decimal import Decimal as _D
        import datetime as _dt
        if isinstance(obj, _D):
            return float(obj)
        if isinstance(obj, (_dt.date, _dt.datetime)):
            return obj.isoformat()
        return super().default(obj)


def _jdumps(obj):
    return json.dumps(obj, cls=_JsonEncoder)


def _month_key(val):
    """TruncMonth on SQLite may return str or datetime — handle both."""
    from datetime import datetime as _dt, date as _d
    if val is None:
        return ''
    if isinstance(val, str):
        return val[:7]
    if isinstance(val, (_dt, _d)):
        return val.strftime('%Y-%m')
    return str(val)[:7]


def _day_key(val):
    """TruncDate on SQLite may return str or datetime — handle both."""
    from datetime import datetime as _dt, date as _d
    if val is None:
        return ''
    if isinstance(val, str):
        return val[:10]
    if isinstance(val, (_dt, _d)):
        return val.strftime('%Y-%m-%d')
    return str(val)[:10]

logger = logging.getLogger(__name__)


# ── Dashboard ─────────────────────────────────────────────

@login_required
def index(request):
    projects = Project.objects.filter(user=request.user).annotate(
        task_count_ann=Count('tasks'),
        hours_total=Sum('tasks__hours'),
    )
    today = date.today()
    month_start = today.replace(day=1)
    from django.db.models import Case, When, Value, IntegerField
    active_tasks = (
        Task.objects
        .filter(
            project__user=request.user,
            status__in=[Task.STATUS_TODO, Task.STATUS_IN_PROGRESS, Task.STATUS_DEFERRED]
        )
        .select_related('project')
        .annotate(
            urgency_order=Case(
                When(due_date__lt=today, then=Value(0)),
                When(due_date=today, then=Value(1)),
                When(due_date__lte=today + timedelta(days=3), then=Value(2)),
                When(due_date__lte=today + timedelta(days=7), then=Value(3)),
                When(due_date__isnull=False, then=Value(4)),
                default=Value(5),
                output_field=IntegerField(),
            )
        )
        .order_by('urgency_order', 'due_date', '-date')[:10]
    )
    recent_tasks = active_tasks

    agg = Task.objects.filter(
        project__user=request.user,
        date__gte=month_start,
    ).aggregate(count=Count('id'), hours=Sum('hours'))

    last_date = (
        Task.objects.filter(project__user=request.user)
        .order_by('-date').values_list('date', flat=True).first()
    )
    gap_days = None
    if last_date:
        delta = (today - last_date).days
        if delta >= 2:
            gap_days = delta

    # ── Dashboard charts ──
    fourteen_days_ago = today - timedelta(days=13)
    daily_qs = (
        Task.objects
        .filter(project__user=request.user, date__gte=fourteen_days_ago)
        .values('date')
        .annotate(count=Count('id'), hours=Sum('hours'))
    )
    daily_map = {str(row['date']): row for row in daily_qs}
    chart_days, chart_counts, chart_hours = [], [], []
    for i in range(14):
        d = fourteen_days_ago + timedelta(days=i)
        k = d.strftime('%Y-%m-%d')
        chart_days.append(d.strftime('%d.%m'))
        row = daily_map.get(k, {})
        chart_counts.append(row.get('count', 0))
        chart_hours.append(float(row.get('hours') or 0))

    # Project distribution by hours this month
    proj_dist = list(
        projects.filter(hours_total__gt=0).order_by('-hours_total')[:6]
    )
    proj_dist_labels = _jdumps([p.name for p in proj_dist])
    proj_dist_hours  = _jdumps([float(p.hours_total or 0) for p in proj_dist])

    done_recent = (
        Task.objects
        .filter(project__user=request.user, status=Task.STATUS_DONE, date__gte=month_start)
        .select_related('project')
        .order_by('-date')[:20]
    )

    # Group active tasks by urgency for sidebar
    active_grouped = {'overdue': [], 'today': [], 'soon': [], 'upcoming': [], 'other': []}
    for t in active_tasks:
        u = t.urgency
        if u in active_grouped:
            active_grouped[u].append(t)
        else:
            active_grouped['other'].append(t)
    active_total = sum(len(v) for v in active_grouped.values())

    return render(request, 'projects/index.html', {
        'projects': projects,
        'recent_tasks': recent_tasks,
        'done_recent': done_recent,
        'active_grouped': active_grouped,
        'active_total': active_total,
        'tasks_count': agg['count'] or 0,
        'tasks_hours': agg['hours'] or Decimal('0'),
        'today': today,
        'gap_days': gap_days,
        'chart_days':        _jdumps(chart_days),
        'chart_counts':      _jdumps(chart_counts),
        'chart_hours':       _jdumps(chart_hours),
        'proj_dist_labels':  proj_dist_labels,
        'proj_dist_hours':   proj_dist_hours,
    })


# ── Projects ──────────────────────────────────────────────

@login_required
@require_POST
def project_create(request):
    name = request.POST.get('name', '').strip()
    if not name:
        messages.error(request, 'Введите название проекта')
        return redirect('index')
    Project.objects.create(
        user=request.user,
        name=name,
        initiator=request.POST.get('initiator', '').strip(),
        description=request.POST.get('description', '').strip(),
    )
    messages.success(request, f'Проект «{name}» создан')
    return redirect('index')


@login_required
@require_POST
def project_delete(request, pk):
    project = get_object_or_404(Project, pk=pk, user=request.user)
    task_count = project.tasks.count()
    name = project.name
    project.delete()
    messages.success(request, f'Проект «{name}» и {task_count} задач удалены')
    return redirect('index')


@login_required
def project_detail(request, pk):
    project = get_object_or_404(Project, pk=pk, user=request.user)
    tasks = project.tasks.select_related('project').all()

    month = request.GET.get('month', '')
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')

    if month:
        tasks = tasks.filter(date__startswith=month)
    if search:
        tasks = tasks.filter(
            Q(task__icontains=search) |
            Q(initiator__icontains=search) |
            Q(basis__icontains=search)
        )
    if status_filter:
        tasks = tasks.filter(status=status_filter)

    total_hours = tasks.aggregate(total=Sum('hours'))['total'] or Decimal('0')
    paginator = Paginator(tasks, getattr(settings, 'TASKS_PER_PAGE', 25))
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'projects/project_detail.html', {
        'project': project,
        'page_obj': page_obj,
        'tasks': page_obj,
        'month': month,
        'search': search,
        'status_filter': status_filter,
        'total_hours': total_hours,
        'status_choices': Task.STATUS_CHOICES,
        'project_total_hours': project.total_hours(),
    })


# ── Tasks ─────────────────────────────────────────────────

def _parse_hours(raw):
    """Parse hours from user input. Accepts '2.5', '2,5', '2h30m', '2ч30м', '2:30'."""
    if not raw:
        return None
    raw = raw.strip().lower()

    # formats: "2ч30м", "2h30m", "2:30"
    import re
    m = re.match(r'^(\d+)[hч][:\s]*(\d+)[mм]?$', raw)
    if m:
        return Decimal(m.group(1)) + Decimal(m.group(2)) / 60

    m = re.match(r'^(\d+):(\d+)$', raw)
    if m:
        return Decimal(m.group(1)) + Decimal(m.group(2)) / 60

    # plain number
    try:
        return Decimal(raw.replace(',', '.'))
    except InvalidOperation:
        raise ValueError(f'Неверный формат: «{raw}». Примеры: 2.5, 2ч30м, 2:30')


@login_required
@require_POST
def task_create(request, project_pk):
    project = get_object_or_404(Project, pk=project_pk, user=request.user)
    task_text = request.POST.get('task', '').strip()
    task_date = request.POST.get('date', '')

    if not task_text or not task_date:
        messages.error(request, 'Заполните дату и описание задачи')
        return redirect('project_detail', pk=project_pk)

    try:
        hours = _parse_hours(request.POST.get('hours', ''))
    except ValueError as e:
        messages.error(request, str(e))
        return redirect('project_detail', pk=project_pk)

    Task.objects.create(
        project=project,
        date=task_date,
        task=task_text,
        status=request.POST.get('status', Task.STATUS_DONE),
        initiator=request.POST.get('initiator', '').strip(),
        hours=hours,
        start_date = request.POST.get('start_date', '').strip() or None,
        due_date   = request.POST.get('due_date', '').strip() or None,
        basis=request.POST.get('basis', '').strip(),
    )
    messages.success(request, 'Задача добавлена')
    return redirect('project_detail', pk=project_pk)


@login_required
@require_POST
def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk, project__user=request.user)
    project_pk = task.project_id
    task.delete()
    messages.success(request, 'Задача удалена')
    return redirect('project_detail', pk=project_pk)


@login_required
def task_edit(request, pk):
    task = get_object_or_404(Task, pk=pk, project__user=request.user)

    if request.method == 'POST':
        task_text = request.POST.get('task', '').strip()
        if not task_text:
            messages.error(request, 'Описание задачи не может быть пустым')
            return render(request, 'projects/task_edit.html', {
                'task': task, 'status_choices': Task.STATUS_CHOICES
            })
        try:
            hours = _parse_hours(request.POST.get('hours', ''))
        except ValueError as e:
            messages.error(request, str(e))
            return render(request, 'projects/task_edit.html', {
                'task': task, 'status_choices': Task.STATUS_CHOICES
            })

        task.date = request.POST.get('date', task.date)
        task.task = task_text
        task.status = request.POST.get('status', task.status)
        task.initiator = request.POST.get('initiator', '').strip()
        task.hours = hours
        task.start_date = request.POST.get('start_date', '').strip() or None
        task.due_date   = request.POST.get('due_date', '').strip() or None
        task.basis = request.POST.get('basis', '').strip()
        task.save()
        messages.success(request, 'Задача обновлена')
        return redirect('project_detail', pk=task.project_id)

    return render(request, 'projects/task_edit.html', {
        'task': task,
        'status_choices': Task.STATUS_CHOICES,
    })


# ── Quick task (navbar shortcut) ──────────────────────────

@login_required
def quick_add(request):
    """Quick-add task from any page via navbar button."""
    projects = Project.objects.filter(user=request.user)
    if request.method == 'POST':
        project_pk = request.POST.get('project_id', '')
        task_text = request.POST.get('task', '').strip()
        task_date = request.POST.get('date', '')
        if not project_pk or not task_text or not task_date:
            messages.error(request, 'Заполните проект, дату и описание')
            return render(request, 'projects/quick_add.html', {
                'projects': projects,
                'status_choices': Task.STATUS_CHOICES,
            })
        project = get_object_or_404(Project, pk=project_pk, user=request.user)
        try:
            hours = _parse_hours(request.POST.get('hours', ''))
        except ValueError as e:
            messages.error(request, str(e))
            return render(request, 'projects/quick_add.html', {
                'projects': projects,
                'status_choices': Task.STATUS_CHOICES,
            })
        Task.objects.create(
            project=project,
            date=task_date,
            task=task_text,
            status=request.POST.get('status', Task.STATUS_DONE),
            initiator=request.POST.get('initiator', '').strip(),
            hours=hours,
        start_date = request.POST.get('start_date', '').strip() or None,
        due_date   = request.POST.get('due_date', '').strip() or None,
            basis=request.POST.get('basis', '').strip(),
        )
        messages.success(request, f'Задача добавлена в «{project.name}»')
        next_url = request.POST.get('next', '/')
        return redirect(next_url)

    return render(request, 'projects/quick_add.html', {
        'projects': projects,
        'status_choices': Task.STATUS_CHOICES,
    })


# ── Report ────────────────────────────────────────────────

@login_required
def report(request):
    projects = Project.objects.filter(user=request.user)
    tasks = Task.objects.filter(
        project__user=request.user
    ).select_related('project')

    today = date.today()
    date_from = request.GET.get('date_from', today.replace(day=1).isoformat())
    date_to = request.GET.get('date_to', today.isoformat())
    project_id = request.GET.get('project', '')
    group_by = request.GET.get('group_by', 'project')

    if date_from:
        tasks = tasks.filter(date__gte=date_from)
    if date_to:
        tasks = tasks.filter(date__lte=date_to)
    if project_id:
        tasks = tasks.filter(project_id=project_id)

    tasks = tasks.order_by('date', 'project__name')

    # Materialise once — avoid double-hit in xlsx path
    tasks_list = list(tasks)
    total_hours = sum(
        (t.hours for t in tasks_list if t.hours), Decimal('0')
    )

    if request.GET.get('export') == 'xlsx':
        return _export_xlsx(tasks_list, date_from, date_to)

    grouped = _group_tasks(tasks_list, group_by)

    return render(request, 'projects/report.html', {
        'projects': projects,
        'tasks': tasks_list,
        'grouped': grouped,
        'group_by': group_by,
        'date_from': date_from,
        'date_to': date_to,
        'project_id': project_id,
        'selected_project': projects.filter(pk=project_id).first() if project_id else None,
        'total': len(tasks_list),
        'total_hours': total_hours,
    })


def _group_tasks(tasks_list, group_by):
    grouped = {}
    if group_by == 'project':
        for t in tasks_list:
            grouped.setdefault(t.project, []).append(t)
    elif group_by == 'date':
        for t in tasks_list:
            grouped.setdefault(t.date, []).append(t)
    else:
        grouped = {'__all__': tasks_list}
    return grouped


def _export_xlsx(tasks_list, date_from, date_to):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Отчёт'

    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='1C1C1A', end_color='1C1C1A', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical='top')
    even_fill = PatternFill(start_color='F5F5F3', end_color='F5F5F3', fill_type='solid')

    headers = ['Дата', 'Проект', 'Задача', 'Статус', 'Инициатор', 'Часы', 'Обоснование']
    widths = [12, 24, 50, 16, 24, 8, 40]

    ws.row_dimensions[1].height = 22
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = center
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    status_map = dict(Task.STATUS_CHOICES)
    total_hours = Decimal('0')

    for ri, task in enumerate(tasks_list, 2):
        h = task.hours or None
        if h:
            total_hours += h
        row = [
            task.date.strftime('%d.%m.%Y'),
            task.project.name,
            task.task,
            status_map.get(task.status, task.status),
            task.initiator or '—',
            float(h) if h else '—',
            task.basis or '—',
        ]
        ws.row_dimensions[ri].height = 40
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = wrap
            if ri % 2 == 0:
                c.fill = even_fill

    # Summary
    sr = len(tasks_list) + 2
    ws.cell(row=sr, column=1, value='Итого').font = Font(bold=True)
    ws.cell(row=sr, column=6, value=float(total_hours)).font = Font(bold=True)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:G{len(tasks_list) + 1}'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    period = f'{date_from}__{date_to}' if (date_from or date_to) else 'all'
    response['Content-Disposition'] = f'attachment; filename="report_{period}.xlsx"'
    wb.save(response)
    return response



@login_required
@require_POST
def project_edit(request, pk):
    project = get_object_or_404(Project, pk=pk, user=request.user)
    name = request.POST.get('name', '').strip()
    if not name:
        messages.error(request, 'Название проекта не может быть пустым')
        return redirect('project_detail', pk=pk)
    project.name = name
    project.initiator = request.POST.get('initiator', '').strip()
    project.description = request.POST.get('description', '').strip()
    project.save()
    messages.success(request, f'Проект «{project.name}» обновлён')
    return redirect('project_detail', pk=pk)


# ── Profile ───────────────────────────────────────────────

@login_required
def profile(request):
    from .models import UserProfile
    profile_obj, _ = UserProfile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        profile_obj.last_name   = request.POST.get('last_name', '').strip()
        profile_obj.first_name  = request.POST.get('first_name', '').strip()
        profile_obj.middle_name = request.POST.get('middle_name', '').strip()
        profile_obj.position    = request.POST.get('position', '').strip()
        profile_obj.department  = request.POST.get('department', '').strip()
        profile_obj.save()
        # also update Django's built-in fields
        request.user.email = request.POST.get('email', '').strip()
        request.user.save(update_fields=['email'])
        messages.success(request, 'Профиль обновлён')
        return redirect('profile')
    return render(request, 'projects/profile.html', {'profile': profile_obj})


# ── Admin Panel ───────────────────────────────────────────


def superuser_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_superuser:
            return HttpResponseForbidden('Доступ запрещён')
        return view_func(request, *args, **kwargs)
    return login_required(wrapper)


@superuser_required
def admin_panel(request):
    users = User.objects.exclude(pk=request.user.pk).annotate(
        project_count=Count('projects'),
        task_count=Count('projects__tasks'),
        hours_total=Sum('projects__tasks__hours'),
    ).order_by('username')

    # global stats
    all_tasks = Task.objects.select_related('project', 'project__user')
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        all_tasks = all_tasks.filter(
            Q(task__icontains=search) |
            Q(project__name__icontains=search) |
            Q(project__user__username__icontains=search)
        )
    if status_filter:
        all_tasks = all_tasks.filter(status=status_filter)
    if user_filter:
        all_tasks = all_tasks.filter(project__user_id=user_filter)
    if date_from:
        all_tasks = all_tasks.filter(date__gte=date_from)
    if date_to:
        all_tasks = all_tasks.filter(date__lte=date_to)

    all_tasks = all_tasks.order_by('-date', '-created_at')

    total_hours = all_tasks.aggregate(t=Sum('hours'))['t'] or Decimal('0')

    paginator = Paginator(all_tasks, 30)
    page_obj = paginator.get_page(request.GET.get('page'))

    all_users = User.objects.exclude(is_superuser=True).order_by('username')

    return render(request, 'projects/admin_panel.html', {
        'users': users,
        'page_obj': page_obj,
        'tasks': page_obj,
        'total_hours': total_hours,
        'total_count': all_tasks.count(),
        'search': search,
        'status_filter': status_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
        'status_choices': Task.STATUS_CHOICES,
        'all_users': all_users,
    })


@superuser_required
def admin_user_detail(request, user_id):
    target_user = get_object_or_404(User, pk=user_id)
    projects = Project.objects.filter(user=target_user).annotate(
        task_count_ann=Count('tasks'),
        hours_total=Sum('tasks__hours'),
    )

    today = date.today()
    month_start = today.replace(day=1)
    tasks_month = Task.objects.filter(
        project__user=target_user, date__gte=month_start
    ).aggregate(count=Count('id'), hours=Sum('hours'))

    recent_tasks = Task.objects.filter(
        project__user=target_user
    ).select_related('project').order_by('-date')[:10]

    total = Task.objects.filter(project__user=target_user).aggregate(
        count=Count('id'), hours=Sum('hours')
    )

    return render(request, 'projects/admin_user_detail.html', {
        'target_user': target_user,
        'projects': projects,
        'recent_tasks': recent_tasks,
        'tasks_month': tasks_month,
        'total': total,
    })


# ── Analytics ─────────────────────────────────────────────


@login_required
def analytics(request):
    today = date.today()

    # ── 1. Monthly dynamics — last 12 months ──
    twelve_months_ago = (today.replace(day=1) - timedelta(days=335)).replace(day=1)

    monthly_qs = (
        Task.objects
        .filter(project__user=request.user, date__gte=twelve_months_ago)
        .values('date__year', 'date__month')
        .annotate(count=Count('id'), hours=Sum('hours'))
        .order_by('date__year', 'date__month')
    )

    # Build full 12-month series (fill gaps with 0)
    months_map = {}
    for row in monthly_qs:
        key = f"{row['date__year']:04d}-{row['date__month']:02d}"
        months_map[key] = {
            'count': row['count'],
            'hours': float(row['hours'] or 0),
        }

    monthly_labels = []
    monthly_counts = []
    monthly_hours = []
    cursor = twelve_months_ago
    while cursor <= today.replace(day=1):
        key = cursor.strftime('%Y-%m')
        label = cursor.strftime('%b %Y')
        monthly_labels.append(label)
        monthly_counts.append(months_map.get(key, {}).get('count', 0))
        monthly_hours.append(round(months_map.get(key, {}).get('hours', 0), 1))
        # next month
        if cursor.month == 12:
            cursor = cursor.replace(year=cursor.year + 1, month=1)
        else:
            cursor = cursor.replace(month=cursor.month + 1)

    # ── 2. Daily activity — last 30 days ──
    thirty_days_ago = today - timedelta(days=29)

    daily_qs = (
        Task.objects
        .filter(project__user=request.user, date__gte=thirty_days_ago)
        .values('date')
        .annotate(count=Count('id'), hours=Sum('hours'))
        .order_by('date')
    )

    daily_map = {}
    for row in daily_qs:
        key = str(row['date'])
        daily_map[key] = {
            'count': row['count'],
            'hours': float(row['hours'] or 0),
        }

    daily_labels = []
    daily_counts = []
    daily_hours = []
    for i in range(30):
        d = thirty_days_ago + timedelta(days=i)
        key = d.strftime('%Y-%m-%d')
        daily_labels.append(d.strftime('%d.%m'))
        daily_counts.append(daily_map.get(key, {}).get('count', 0))
        daily_hours.append(round(daily_map.get(key, {}).get('hours', 0), 1))

    # ── 3. Status breakdown (all time) ──
    status_qs = (
        Task.objects
        .filter(project__user=request.user)
        .values('status')
        .annotate(count=Count('id'))
    )
    status_map_labels = dict(Task.STATUS_CHOICES)
    status_labels = []
    status_counts = []
    for row in status_qs:
        status_labels.append(status_map_labels.get(row['status'], row['status']))
        status_counts.append(row['count'])

    # ── 4. Top projects by hours ──
    top_projects_qs = (
        Project.objects
        .filter(user=request.user)
        .annotate(hours_total=Sum('tasks__hours'), task_count=Count('tasks'))
        .filter(hours_total__gt=0)
        .order_by('-hours_total')[:8]
    )
    proj_labels = [p.name for p in top_projects_qs]
    proj_hours = [float(p.hours_total or 0) for p in top_projects_qs]

    # ── Summary stats ──
    total_all = Task.objects.filter(project__user=request.user).aggregate(
        count=Count('id'), hours=Sum('hours')
    )
    this_month = Task.objects.filter(
        project__user=request.user,
        date__gte=today.replace(day=1)
    ).aggregate(count=Count('id'), hours=Sum('hours'))
    last_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
    last_month_end = today.replace(day=1) - timedelta(days=1)
    last_month = Task.objects.filter(
        project__user=request.user,
        date__gte=last_month_start,
        date__lte=last_month_end,
    ).aggregate(count=Count('id'), hours=Sum('hours'))

    return render(request, 'projects/analytics.html', {
        'monthly_labels': _jdumps(monthly_labels),
        'monthly_counts': _jdumps(monthly_counts),
        'monthly_hours':  json.dumps(monthly_hours),
        'daily_labels':   json.dumps(daily_labels),
        'daily_counts':   json.dumps(daily_counts),
        'daily_hours':    json.dumps(daily_hours),
        'status_labels':  json.dumps(status_labels),
        'status_counts':  json.dumps(status_counts),
        'proj_labels':    json.dumps(proj_labels),
        'proj_hours':     json.dumps(proj_hours),
        'total_all':      total_all,
        'this_month':     this_month,
        'last_month':     last_month,
        'today':          today,
    })
